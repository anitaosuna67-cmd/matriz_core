import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

def crear_excel_estilizado(filename, title, subtitle, headers, data, sheet_name="Datos"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.views.sheetView[0].showGridLines = True

    # Paleta de colores analítica
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    derecha_fill = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")
    izquierda_fill = PatternFill(start_color="FAF2F2", end_color="FAF2F2", fill_type="solid")
    neutral_fill = PatternFill(start_color="FDF5E6", end_color="FDF5E6", fill_type="solid")

    font_title = Font(name="Arial", size=16, bold=True, color="1F4E78")
    font_subtitle = Font(name="Arial", size=11, italic=True, color="595959")
    font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Arial", size=10, color="000000")
    font_body_bold = Font(name="Arial", size=10, bold=True, color="000000")

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'), top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))

    ws["A1"] = title
    ws["A1"].font = font_title
    ws["A2"] = subtitle
    ws["A2"].font = font_subtitle

    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_idx, value=text)
        cell.font = font_header
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border

    row_idx = 6
    for fila in data:
        faccion = str(fila[4]) if len(fila) > 4 else ""
        for col_idx, valor in enumerate(fila, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.font = font_body
            cell.border = thin_border
            
            if "DERECHA" in faccion:
                cell.fill = derecha_fill
            elif "IZQUIERDA" in faccion:
                cell.fill = izquierda_fill
            else:
                cell.fill = neutral_fill
                
            if col_idx in [1, 2, 3, 4, 5, 7] and len(headers) >= col_idx:
                if col_idx == 5 or (len(headers) == 6 and col_idx == 5):
                    cell.font = font_body_bold
                cell.alignment = align_center
            else:
                cell.alignment = align_left
        row_idx += 1

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row > 3 and cell.value:
                lines = str(cell.value).split('\n')
                for line in lines:
                    if len(line) > max_len: max_len = len(line)
        ws.column_dimensions[col_letter].width = min(max_len + 3, 42) if max_len > 0 else 15

    # Crear directorios automáticamente si no existen
    os.makedirs(os.path.dirname(filename), exist_ok=True)
    wb.save(filename)
    print(f"✅ Generado con éxito: {filename}")

if __name__ == "__main__":
    print("Iniciando generación masiva de archivos del ecosistema...")
    
    base_dir = "data/processed/"

    # --- DATOS VECTOR 1: ESOTÉRICO ---
    v1_h = ["Sistema", "Entidad / Objeto", "Clasificación Interna", "Operación Matriz", "Facción Alineada", "Descripción Estructural", "Raíz Asociada"]
    v1_d = [
        ("Enoquiano (John Dee)", "Mesa Santa (Holy Table)", "Artefacto de Conexión", "Control / Geopolítica", "DERECHA (Control)", "Sistema de caracteres diseñado para centralizar la recepción de instrucciones imperiales bajo Isabel I.", "Yamin (Derecha)"),
        ("Enoquiano (John Dee)", "Nalvage", "Entidad de Orden", "Estructuración Territorial", "DERECHA (Control)", "Regula los límites geográficos y la división del mundo en cuadrantes rígidos de vigilancia.", "Beyn (Separación)"),
        ("Salomónico (Goetia)", "Rey Asmodeo", "Jerarquía Suprema", "Disolución Moral / Caos", "IZQUIERDA (Caos)", "Taxonomía orientada a la ruptura de acuerdos y la introducción del desorden emocional.", "Smol (Izquierda)"),
        ("Salomónico (Goetia)", "Duque Bune", "Rango Duque", "Mutación Material / Riqueza", "IZQUIERDA (Caos)", "Invocado para alterar la acumulación de bienes mediante la manipulación de la materia.", "Yada (Percepción)"),
        ("Cábala Zohárica", "Sefirot de la Derecha (Chesed/Chochmah)", "Emanación de Expansión", "Arquitectura del Sistema", "DERECHA (Control)", "Fuerzas de cohesión, orden absoluto y estructura jerárquica de la luz emanada.", "Yamin (Derecha)"),
        ("Cábala Zohárica", "Sitra Achra (El Otro Lado)", "Emanación del Caos / Qliphoth", "Fragmentación de Límites", "IZQUIERDA (Caos)", "Estructuras de cáscaras que rompen los contenedores originales introduciendo la entropía celular.", "Smol (Izquierda)"),
        ("Gnosticismo Antiguo", "Demiurgo (Yaldabaoth)", "Gobernante de la Materia", "Tiranía Piramidal", "DERECHA (Control)", "Construcción de una prisión material basada en leyes rígidas y vigilancia aróntica de las almas.", "Yamin (Derecha)"),
        ("Misterios de Eleusis", "Iniciación de Perséfone", "Rito de Transgresión / Visión", "Alteración Perceptiva", "IZQUIERDA (Caos)", "Quiebre de los límites de la percepción ordinaria para asimilar la mutación del ciclo de la muerte.", "Yada (Percepción)"),
        ("Hermetismo Egipcio", "Tablas de Esmeralda", "Leyes de Correspondencia", "Definición de Filtros", "DERECHA (Control)", "Protocolo estricto que dicta las reglas inmutables de cómo lo de arriba domina a lo de abajo.", "Beyn (Separación)")
    ]
    crear_excel_estilizado(base_dir + "Vector_1_Esoterico.xlsx", "VECTOR ESOTÉRICO & MITOLÓGICO (CAPA 1)", "Sistematización de manuales operativos, jerarquías y herramientas de conexión.", v1_h, v1_d, "Datos Esotéricos")

    # --- DATOS VECTOR 2: GEOPOLÍTICO ---
    v2_h = ["Entidad / Corporación", "Época", "Área de Influencia", "Mecanismo de Control", "Archivo / Patente Referencia", "Facción Alineada", "Impacto Estructural"]
    v2_d = [
        ("Compañía Británica de las Indias Orientales", "1600 - 1874", "Comercio Global", "Monopolio corporativo con ejército", "Archivos del Parlamento Británico", "DERECHA (Control)", "Primera mega-corporación; privatización del dominio territorial y financiero."),
        ("Standard Oil", "1870 - 1911", "Energía e Infraestructura", "Integración horizontal y vertical", "Caso Corte Suprema de EE.UU. (1911)", "DERECHA (Control)", "Monopolio energético global que condicionó el desarrollo industrial independiente."),
        ("DARPA", "1958 - Presente", "Vigilancia y Tecnología", "Financiamiento de redes e internet", "Presupuestos del Pentágono (Desclasificados)", "DERECHA (Control)", "Creación de ARPANET con fines de control de comunicaciones y vigilancia militar."),
        ("BlackRock / Vanguard", "1988 - Presente", "Flujo de Capitales", "Concentración de acciones globales", "Informes SEC (Formularios 13F)", "DERECHA (Control)", "Centralización de la propiedad accionaria internacional, homogeneizando políticas globales."),
        ("Imperio Romano", "27 a.C. - 476 d.C.", "Centralización Territorial", "Fiscus Judaicus / Derecho Romano", "Códices de Justiniano", "DERECHA (Control)", "Institucionalización del modelo de vasallaje fiscal y codificación de la propiedad privada piramidal."),
        ("Orden del Temple", "1119 - 1312", "Banca Transnacional", "Red de pagarés y crédito de tránsito", "Bulas Papales (Omne Datum Optimum)", "DERECHA (Control)", "Primer sistema financiero descentralizado de la corona pero centralizado en el control del oro europeo."),
        ("Banco de Pagos Internacionales (BIS)", "1930 - Presente", "Liquidación Financiera", "Banca central de bancos centrales", "Tratado de La Haya", "DERECHA (Control)", "Coordinación inmune a leyes locales que dicta las tasas y restricciones de liquidez del planeta."),
        ("Sistema de Crédito Social Chino", "2014 - Presente", "Control Conductual", "Algoritmos de vigilancia masiva e IA", "Directivas del Consejo de Estado Chino", "DERECHA (Control)", "Fusión total del Estado y la tecnología para anular la disidencia mediante castigo financiero inmediato.")
    ]
    crear_excel_estilizado(base_dir + "Vector_2_Geopolitico.xlsx", "VECTOR GEOPOLÍTICO & FINANCIERO (CAPA 2)", "Mapeo de corporaciones, monopolios y archivos institucionales ejecutores.", v2_h, v2_d, "Estructuras de Control")

    # --- DATOS VECTOR 3: BIOTECNOLÓGICO ---
    v3_h = ["Avance / Plataforma", "Campo Científico", "Mecanismo de Operación", "Registro / Patente Referencia", "Facción Alineada", "Impacto Estructural en la Matriz"]
    v3_d = [
        ("CRISPR-Cas9 / Edición Genética", "Biotecnología Aplicada", "Modificación de la cadena de ADN", "Patentes Broad Institute", "IZQUIERDA (Caos)", "Ruptura de los límites biológicos heredados; capacidad de rediseñar especies fuera del orden."),
        ("Neuralink / Interfaz Cerebro", "Neurotecnología", "Inyección de electrodos neuronales", "Registros de ensayos FDA", "IZQUIERDA (Caos)", "Disolución de la barrera analógica del pensamiento humano, conectando la carne a redes sintéticas."),
        ("EctoLife / Úteros Artificiales", "Ingeniería Reproductiva", "Gestación extracorpórea por IA", "Modelos de bio-incubadoras", "IZQUIERDA (Caos)", "Desconexión absoluta del diseño reproductivo natural y del linaje materno biológico."),
        ("Biología Sintética / ARNm", "Genómica Sintética", "Programación de instrucciones celulares", "Registros de patentes farmacéuticas", "IZQUIERDA (Caos)", "Conversión del cuerpo en un software reescribible sujeto a actualizaciones externas."),
        ("Quimeras Humano-Animal", "Ingeniería de Tejidos", "Inyección de células madre en blastocistos", "Ensayos del Salk Institute", "IZQUIERDA (Caos)", "Disolución de las fronteras taxonómicas específicas entre especies del diseño original."),
        ("Carne de Cultivo Celular", "Bioregeneración Alimentaria", "Proliferación in vitro de miocitos", "Aprobaciones FDA / USDA", "IZQUIERDA (Caos)", "Modificación de la cadena de asimilación energética celular, independizándola de la tierra."),
        ("Nanotecnología Lipídica", "Nanomedicina", "Vectores autoensamblados intracelulares", "Patentes de bionanotecnología", "IZQUIERDA (Caos)", "Introducción de componentes sintéticos a nivel subcelular para interactuar con frecuencias externas.")
    ]
    crear_excel_estilizado(base_dir + "Vector_3_Biotecnologico.xlsx", "VECTOR BIOTECNOLÓGICO (CAPA 3)", "Monitoreo de ingeniería genética, transhumanismo e interfaces de disolución.", v3_h, v3_d, "Alteraciones Biológicas")

    # --- DATOS VECTOR 4: ASTRO-TEMPORAL ---
    v4_h = ["Ciclo / Marcador Celeste", "Coordenada Astrocartográfica", "Evento Histórico Correlacionado", "Ventana de Ejecución", "Facción Alineada", "Impacto Estructural"]
    v4_d = [
        ("Alineación de Regulus", "Línea Zenit - Washington / Londres", "Fundación de infraestructuras imperiales", "Ciclos de 72 años", "DERECHA (Control)", "Fijación del poder autoritario y consolidación de la arquitectura piramidal."),
        ("Retorno de Plutón", "Líneas de Medio Cielo (MC)", "Caída de Babilonia / Crisis EE.UU.", "Ciclo de 248 años", "IZQUIERDA (Caos)", "Colapso del orden material establecido y mutación violenta del sistema."),
        ("Ciclos Saros (Eclipses)", "Nodos de intersección variables", "Fracturas de tratados y caídas dinásticas", "18 años y 11 días", "Ambas Facciones", "Puntos de inflexión donde se introduce ruido en el código o se resetean acuerdos."),
        ("Conjunción Júpiter-Saturno", "Líneas de Ascendente (AC)", "Cambios de paradigma económico global", "Ciclos de 20 / 200 años", "DERECHA (Control)", "Rediseño de las herramientas de control poblacional; transiciones de eras elementales."),
        ("Mínimo de Maunder (Ciclo Solar)", "Afectación heliosférica global", "Pequeña Edad de Hielo / Hambrunas", "Ciclos de 11 / 400 años", "IZQUIERDA (Caos)", "Estrés térmico y biológico que desestabiliza las cosechas, induciendo caos social y revueltas."),
        ("Tránsito de Urano por Tauro", "Líneas angulares sobre centros bancarios", "Gran Depresión (1930s) / Criptoactivos", "Ciclo de 84 años", "IZQUIERDA (Caos)", "Disrupción radical de las estructuras de valor material y aceleración tecnológica imprevista."),
        ("Ciclo Metónico de la Luna", "Intersección de calendarios litúrgicos", "Decretos papales y edictos imperiales", "Ciclo de 19 años", "DERECHA (Control)", "Sincronización exacta del tiempo sagrado con el tiempo civil para atrapar la psique de las masas.")
    ]
    crear_excel_estilizado(base_dir + "Vector_4_Astro.xlsx", "VECTOR ASTRO-TEMPORAL (CAPA 4)", "Cálculo de ventanas de oportunidad, astrocartografía histórica y ciclos cíclicos.", v4_h, v4_d, "Reloj del Sistema")

    print("\n All files successfully written directly to data/processed/")
