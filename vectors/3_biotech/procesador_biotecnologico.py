import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

def generar_excel_biotecnologico():
    ruta_guardado = "../../data/processed/Vector_3_Biotecnologico.xlsx"
    print("Iniciando generación del archivo Excel para el Vector Biotecnológico...")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Alteraciones Biológicas"
    ws.views.sheetView[0].showGridLines = True

    # Estilos (Paleta Azul Noche / Tinte Rojo Sutil para la Izquierda/Caos)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    izquierda_fill = PatternFill(start_color="FAF2F2", end_color="FAF2F2", fill_type="solid")

    font_title = Font(name="Arial", size=16, bold=True, color="1F4E78")
    font_subtitle = Font(name="Arial", size=11, italic=True, color="595959")
    font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Arial", size=10, color="000000")
    font_body_bold = Font(name="Arial", size=10, bold=True, color="000000")

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'), top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))

    # Títulos
    ws["A1"] = "VECTOR BIOTECNOLÓGICO (CAPA 3)"
    ws["A2"] = "Monitoreo de ingeniería genética, transhumanismo e interfaces de disolución del diseño natural."
    ws["A1"].font = font_title
    ws["A2"].font = font_subtitle

    # Encabezados
    headers = ["Avance / Plataforma", "Campo Científico", "Mecanismo de Operación", "Registro / Patente Referencia", "Facción Alineada", "Impacto Estructural en la Matriz"]
    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_idx, value=text)
        cell.font = font_header
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border

    # Datos (Enfocados en la "Izquierda" / Disolución y Caos Biológico)
    datos = [
        ("CRISPR-Cas9 / Edición Genética", "Biotecnología Aplicada", "Modificación directa de la cadena de ADN", "Patentes Broad Institute / UC Berkeley", "IZQUIERDA (Caos)", "Ruptura irreversible de los límites biológicos heredados; capacidad de rediseñar las especies fuera del orden natural."),
        ("Neuralink / Interfaces Cerebro-Máquina", "Neurotecnología", "Inyección de electrodos para lectura/escritura neuronal", "Registros de ensayos clínicos FDA", "IZQUIERDA (Caos)", "Disolución de la barrera analógica del pensamiento humano, permitiendo la hibridación directa con inteligencias sintéticas."),
        ("EctoLife / Úteros Artificiales", "Ingeniería de la Reproducción", "Gestación extracorpórea controlada por IA", "Modelos conceptuales y registros de bio-incubadoras", "IZQUIERDA (Caos)", "Desconexión absoluta del diseño reproductivo natural y del linaje biológico materno tradicional."),
        ("Biología Sintética / ARNm Avanzado", "Genómica Sintética", "Programación de instrucciones celulares exógenas", "Registros de patentes farmacéuticas globales", "IZQUIERDA (Caos)", "Conversión del cuerpo biológico en un software reescribible, sujeto a actualizaciones y dependencias externas.")
    ]

    row_idx = 6
    for fila in datos:
        for col_idx, valor in enumerate(fila, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.font = font_body
            cell.border = thin_border
            cell.fill = izquierda_fill
            
            if col_idx in [1, 2, 5]:
                cell.alignment = align_center
                if col_idx == 5: cell.font = font_body_bold
            else:
                cell.alignment = align_left
        row_idx += 1

    # Auto-ajuste de columnas
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row > 3 and cell.value:
                lines = str(cell.value).split('\n')
                for line in lines:
                    if len(line) > max_len: max_len = len(line)
        ws.column_dimensions[col_letter].width = min(max_len + 3, 45) if max_len > 0 else 15

    os.makedirs(os.path.dirname(ruta_guardado), exist_ok=True)
    wb.save(ruta_guardado)
    print(f"¡Éxito! Archivo generado en: {ruta_guardado}")

if __name__ == "__main__":
    generar_excel_biotecnologico()
