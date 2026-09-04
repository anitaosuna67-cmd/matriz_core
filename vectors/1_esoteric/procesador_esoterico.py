import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

def generar_excel_esoterico():
    ruta_guardado = "../../data/processed/Vector_1_Esoterico.xlsx"
    print("Iniciando generación del archivo Excel para el Vector Esotérico...")

    # Crear libro de trabajo
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Datos Esotéricos"
    ws.views.sheetView[0].showGridLines = True

    # Estilos (Paleta Azul Noche / Gris Técnico)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    derecha_fill = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")
    izquierda_fill = PatternFill(start_color="FAF2F2", end_color="FAF2F2", fill_type="solid")

    font_title = Font(name="Arial", size=16, bold=True, color="1F4E78")
    font_subtitle = Font(name="Arial", size=11, italic=True, color="595959")
    font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Arial", size=10, color="000000")
    font_body_bold = Font(name="Arial", size=10, bold=True, color="000000")

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # Títulos decorativos de cabecera
    ws["A1"] = "VECTOR ESOTÉRICO & MITOLÓGICO (CAPA 1)"
    ws["A1"].font = font_title
    ws["A2"] = "Sistematización de manuales operativos, jerarquías y herramientas de conexión de las élites."
    ws["A2"].font = font_subtitle

    # Encabezados de la tabla (Fila 5)
    headers = ["Sistema", "Entidad / Objeto", "Clasificación Interna", "Operación Matriz", "Facción Alineada", "Descripción Estructural", "Raíz Asociada"]
    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_idx, value=text)
        cell.font = font_header
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border

    # Datos duros
    datos = [
        ("Enoquiano (John Dee)", "Mesa Santa (Holy Table)", "Artefacto de Conexión", "Control / Geopolítica", "DERECHA (Control)", "Sistema de caracteres y cuadrados mágicos diseñado para centralizar la recepción de instrucciones. Utilizado para fundamentar la legitimidad y expansión del Imperio Británico bajo Isabel I.", "Yamin (Derecha/Autoridad)"),
        ("Enoquiano (John Dee)", "Nalvage", "Entidad de Orden", "Estructuración Territorial", "DERECHA (Control)", "Entidad que dictó las tablas de las Torres de Guardia. Regula los límites geográficos y la división del mundo en cuadrantes rígidos de vigilancia.", "Beyn (Separación/Límites)"),
        ("Salomónico (Goetia)", "Rey Asmodeo", "Rango Rey (Jerarquía Suprema)", "Disolución Moral / Caos", "IZQUIERDA (Caos)", "Taxonomía orientada a la ruptura de los acuerdos matrimoniales, la introducción del desorden emocional y la alteración de los juicios naturales.", "Smol (Izquierda/Entropía)"),
        ("Salomónico (Goetia)", "Duque Bune", "Rango Duque", "Mutación Material / Riqueza", "IZQUIERDA (Caos)", "Entidad invocada para alterar la acumulación de bienes materiales y cambiar la posición social mediante la manipulación del entorno físico.", "Yada (Percepción/Alteración)")
    ]

    # Inyección de datos y estilos por fila
    row_idx = 6
    for fila in datos:
        for col_idx, valor in enumerate(fila, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.font = font_body
            cell.border = thin_border
            
            # Alineación diferencial
            if col_idx in [1, 2, 3, 4, 5, 7]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left
                
            # Formato condicional según la facción
            if "DERECHA" in fila[4]:
                cell.fill = derecha_fill
                if col_idx == 5: cell.font = font_body_bold
            else:
                cell.fill = izquierda_fill
                if col_idx == 5: cell.font = font_body_bold
                
        row_idx += 1

    # Auto-ajuste inteligente de columnas
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row > 3 and cell.value:
                lines = str(cell.value).split('\n')
                for line in lines:
                    if len(line) > max_len: max_len = len(line)
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50) if max_len > 0 else 15

    # Guardar
    os.makedirs(os.path.dirname(ruta_guardado), exist_ok=True)
    wb.save(ruta_guardado)
    print(f"¡Éxito! Archivo de Excel generado en: {ruta_guardado}")

if __name__ == "__main__":
    generar_excel_esoterico()
