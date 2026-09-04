import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

def generar_excel_geopolitico():
    ruta_guardado = "../../data/processed/Vector_2_Geopolitico.xlsx"
    print("Iniciando generación del archivo Excel para el Vector Geopolítico...")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estructuras de Control"
    ws.views.sheetView[0].showGridLines = True

    # Estilos (Paleta Azul Noche / Gris Técnico)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    derecha_fill = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")

    font_title = Font(name="Arial", size=16, bold=True, color="1F4E78")
    font_subtitle = Font(name="Arial", size=11, italic=True, color="595959")
    font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Arial", size=10, color="000000")
    font_body_bold = Font(name="Arial", size=10, bold=True, color="000000")

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'), top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))

    # Títulos
    ws["A1"] = "VECTOR GEOPOLÍTICO & FINANCIERO (CAPA 2)"
    ws["A1"].font = font_title
    ws["A2"] = "Mapeo de corporaciones, monopolios y archivos institucionales ejecutores de la matriz material."
    ws["A2"].font = font_subtitle

    # Encabezados
    headers = ["Entidad / Corporación", "Época", "Área de Influencia", "Mecanismo de Control", "Archivo / Patente Referencia", "Facción Alineada", "Impacto Estructural"]
    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_idx, value=text)
        cell.font = font_header
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border

    # Datos (Enfocados en la "Derecha" / Centralización)
    datos = [
        ("Compañía Británica de las Indias Orientales", "1600 - 1874", "Comercio Global y Militarización", "Monopolio corporativo con ejército privado", "Archivos del Parlamento Británico (Actas de Fletamiento)", "DERECHA (Control)", "Primera mega-corporación moderna; privatización de la conquista territorial y centralización de rutas comerciales."),
        ("Standard Oil", "1870 - 1911", "Energía e Infraestructura", "Integración horizontal y vertical", "Corte Suprema de EE.UU. (Caso Standard Oil Co. de Nueva Jersey)", "DERECHA (Control)", "Establecimiento del modelo de monopolio energético global, condicionando el desarrollo tecnológico independiente."),
        ("DARPA / Complejo Militar-Industrial", "1958 - Presente", "Vigilancia y Desarrollo Tecnológico", "Financiamiento de redes neuronales e internet", "Registros de presupuestos del Pentágono (Desclasificados)", "DERECHA (Control)", "Creación de la arquitectura base de la red moderna (ARPANET) con fines iniciales de control de comunicaciones y vigilancia militar."),
        ("BlackRock / Vanguard Group", "1988 - Presente", "Flujo de Capitales (Gestión de Activos)", "Concentración de acciones corporativas", "Informes SEC (Formularios 13F)", "DERECHA (Control)", "Centralización sin precedentes de la propiedad accionaria global, homogeneizando las políticas corporativas internacionales a través del poder de voto.")
    ]

    row_idx = 6
    for fila in datos:
        for col_idx, valor in enumerate(fila, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.font = font_body
            cell.border = thin_border
            cell.fill = derecha_fill # Todos son de control en esta muestra
            
            if col_idx in [1, 2, 6]:
                cell.alignment = align_center
                if col_idx == 6: cell.font = font_body_bold
            else:
                cell.alignment = align_left
        row_idx += 1

    # Auto-ajuste de columnas
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row > 3 and cell.value:
                lines = str(cell.value).split('\n')
                for line in lines:
                    if len(line) > max_len: max_len = len(line)
        ws.column_dimensions[col_letter].width = min(max_len + 3, 45) if max_len > 0 else 15

    os.makedirs(os.path.dirname(ruta_guardado), exist_ok=True)
    wb.save(ruta_guardado)
    print(f"¡Éxito! Archivo generado en: {ruta_guardado}")

if __name__ == "__main__":
    generar_excel_geopolitico()
