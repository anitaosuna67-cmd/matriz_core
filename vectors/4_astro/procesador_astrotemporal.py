import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

def generar_excel_astro():
    ruta_guardado = "../../data/processed/Vector_4_Astro.xlsx"
    print("Iniciando generación del archivo Excel para el Reloj Astro-Temporal...")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reloj del Sistema"
    ws.views.sheetView[0].showGridLines = True

    # Estilos (Paleta Azul Noche y Tonos de Facciones)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    derecha_fill = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")
    izquierda_fill = PatternFill(start_color="FAF2F2", end_color="FAF2F2", fill_type="solid")

    font_title = Font(name="Arial", size=16, bold=True, color="1F4E78")
    font_subtitle = Font(name="Arial", size=11, italic=True, color="595959")
    font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Arial", size=10, color="000000")
    font_body_bold = Font(name="Arial", size=10, bold=True, color="000000")

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'), top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))

    # Títulos
    ws["A1"] = "VECTOR ASTRO-TEMPORAL (CAPA 4)"
    ws["A2"] = "Cálculo de ventanas de oportunidad, astrocartografía histórica y ciclos de ejecución matricial."
    ws["A1"].font = font_title
    ws["A2"].font = font_subtitle

    # Encabezados
    headers = ["Ciclo / Marcador Celeste", "Coordenada Astrocartográfica", "Evento Histórico Correlacionado", "Ventana de Ejecución", "Facción Alineada", "Impacto Estructural"]
    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_idx, value=text)
        cell.font = font_header
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border

    # Datos
    datos = [
        ("Alineación de Regulus (Corazón del León)", "Línea Zenit - Washington D.C. / Londres", "Fundación de infraestructuras imperiales y bancos centrales", "Ciclos de 72 años (Precesión)", "DERECHA (Control)", "Fijación del poder autoritario y consolidación de la arquitectura piramidal de las potencias hegemónicas."),
        ("Retorno de Plutón", "Líneas de Medio Cielo (MC) sobre potencias", "Caída de Babilonia / Revolución Francesa / Crisis EE.UU. 2020s", "Ciclo de 248 años", "IZQUIERDA (Caos)", "Colapso del orden material establecido, disolución de fronteras financieras y mutación violenta del sistema."),
        ("Ciclos Saros (Eclipses Solares y Lunares)", "Nodos de intersección (Latitud/Longitud variable)", "Asesinatos de figuras clave y fracturas de tratados internacionales", "Ventanas de 18 años y 11 días", "Ambas Facciones", "Puntos de inflexión donde se introducen alteraciones en el código o se 'resetean' acuerdos geopolíticos."),
        ("Conjunción Júpiter-Saturno (Gran Mutación)", "Líneas de Ascendente (AC) sobre epicentros tecnológicos", "Cambios de paradigma económico (de la Tierra al Aire)", "Ciclos de 20 y 200 años", "DERECHA (Control)", "Rediseño de las herramientas de control poblacional; transición del control de la tierra al control de la información (digital).")
    ]

    row_idx = 6
    for fila in datos:
        for col_idx, valor in enumerate(fila, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.font = font_body
            cell.border = thin_border
            
            # Formato condicional
            if "DERECHA" in fila[4]:
                cell.fill = derecha_fill
            elif "IZQUIERDA" in fila[4]:
                cell.fill = izquierda_fill
            else:
                cell.fill = PatternFill(start_color="FDF5E6", end_color="FDF5E6", fill_type="solid") # Tinte arena para neutrales
                
            if col_idx in [1, 2, 4, 5]:
                cell.alignment = align_center
                if col_idx == 5: cell.font = font_body_bold
            else:
                cell.alignment = align_left
        row_idx += 1

    # Auto-ajuste de columnas
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row > 3 and cell.value:
                lines = str(cell.value).split('\n')
                for line in lines:
                    if len(line) > max_len: max_len = len(line)
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40) if max_len > 0 else 15

    os.makedirs(os.path.dirname(ruta_guardado), exist_ok=True)
    wb.save(ruta_guardado)
    print(f"¡Éxito! Archivo generado en: {ruta_guardado}")

if __name__ == "__main__":
    generar_excel_astro()
